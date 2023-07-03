import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, Font, PatternFill, Border, Side

class ColumnsConvert:
    def __init__(self, writer, sheet):
        # self.file_path = file_path
        # self.wb = openpyxl.load_workbook(file_path)
        # self.ws = self.wb.active
        self.wb = writer.book
        self.ws = writer.sheets[sheet]
        self.format = '"$"#,##0_);("$"#,##0)'
        self.style = ''

    def add_format_currency(self, columns):
        self.style = NamedStyle(name='Currency')
        self.style.number_format = self.format
        # Aplicar o estilo à coluna A
        for row in range(1, self.ws.max_row + 1):

            for column in columns:
                # Obter a célula da coluna A na linha atual
                cell = self.ws.cell(row=row, column=column)
                # Aplicar o estilo à célula
                cell.style = self.style

    def salvar(self, name):
        self.wb.save(name)

    def auto_size_column(self, col_name):
        max_width = 0
        for cell in self.ws[col_name]:
            # Calcular a largura do texto na célula usando a fonte padrão
            text_width = len(str(cell.value)) * 1.2
            # Atualizar a largura máxima se necessário
            if text_width > max_width:
                max_width = text_width
        # Definir a largura da coluna A como a largura máxima arredondada para cima
        self.ws.column_dimensions[col_name].width = int(max_width + 0.5)

    def add_width_columns(self):
        for col in range(1, self.ws.max_column + 1):
            # Obter o nome da coluna em letra
            col_name = get_column_letter(col)
            self.auto_size_column(col_name)

    def create_style_to_header(self):
        header = NamedStyle(name="header")

        # Definir a fonte do estilo como Arial, negrito e preto
        header.font = Font(name="Arial", bold=True, color="000000")

        # Definir o preenchimento do estilo como laranja sólido
        header.fill = PatternFill(fill_type="solid", start_color="FFA500")

        # Definir a borda do estilo como fina e preta
        side = Side(border_style="thin", color="000000")
        header.border = Border(left=side, right=side, top=side, bottom=side)

        # Registrar o estilo no workbook
        self.wb.add_named_style(header)

        for col in range(1, self.ws.max_column + 1):
            # Obter o nome da coluna em letra
            col_name = get_column_letter(col)
            # Obter a célula da primeira linha na coluna atual
            cell = self.ws[f'{col_name}1']
            # Aplicar o estilo à célula
            cell.style = header


# convert = ColumnsConvert('Data_produtos.xlsx')
# convert.add_format_currency([2, 3, 5])
# convert.add_width_columns()
# convert.create_style_to_header()
# convert.salvar()
